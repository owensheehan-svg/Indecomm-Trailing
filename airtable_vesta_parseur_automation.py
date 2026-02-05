#!/usr/bin/env python3
"""
Airtable + Vesta + Parseur to Indecomm Excel Generator
Fetches loan data from multiple sources and generates Indecomm template
"""

import os
import sys
import json
import requests
from datetime import datetime
from typing import Dict, List, Optional
import openpyxl
from openpyxl import load_workbook
import time
from geopy.geocoders import Nominatim
from geopy.exc import GeocoderTimedOut, GeocoderServiceError

# Configuration
AIRTABLE_TOKEN = os.environ.get('AIRTABLE_TOKEN')
AIRTABLE_BASE_ID = 'appgBl5EHB3qFtOPl'
AIRTABLE_TABLE_NAME = 'Post-Close Tracker'
AIRTABLE_VIEW = 'viww1LG42sIrTNGFc'

VESTA_API_KEY = 'vA9kkt@dTcyRtm@yxs.vWfbJKnw6NWNy'
VESTA_BASE_URL = 'https://multiply.beta.vesta.com/api'
VESTA_VERSION = '_26_1'

PARSEUR_API_KEY = 'sk_ea922e2b4f9020f29f41917a28c59e522cf4b91fdb13cff217678dfcb514eabd'
PARSEUR_MAILBOX_NAME = 'cherubic-select-alligator'

TEMPLATE_PATH = 'Funded_File_Template.xlsx'

class IndecommAutomation:
    def __init__(self):
        self.errors = []
        self.warnings = []
        self.geolocator = Nominatim(user_agent="indecomm_automation")
        
    def log_error(self, loan_number: str, field: str, message: str):
        """Log an error for reporting"""
        self.errors.append(f"Loan {loan_number} - {field}: {message}")
        print(f"ERROR: Loan {loan_number} - {field}: {message}")
        
    def log_warning(self, loan_number: str, field: str, message: str):
        """Log a warning for reporting"""
        self.warnings.append(f"Loan {loan_number} - {field}: {message}")
        print(f"WARNING: Loan {loan_number} - {field}: {message}")
    
    def fetch_airtable_records(self) -> List[Dict]:
        """Fetch records from Airtable view"""
        url = f'https://api.airtable.com/v0/{AIRTABLE_BASE_ID}/{AIRTABLE_TABLE_NAME}'
        headers = {
            'Authorization': f'Bearer {AIRTABLE_TOKEN}',
            'Content-Type': 'application/json'
        }
        
        all_records = []
        offset = None
        
        print(f"Fetching records from Airtable view: {AIRTABLE_VIEW}")
        
        while True:
            params = {'view': AIRTABLE_VIEW}
            if offset:
                params['offset'] = offset
                
            response = requests.get(url, headers=headers, params=params)
            
            if response.status_code != 200:
                raise Exception(f"Airtable API error: {response.status_code} - {response.text}")
            
            data = response.json()
            all_records.extend(data.get('records', []))
            
            offset = data.get('offset')
            if not offset:
                break
                
            time.sleep(0.2)
        
        print(f"Found {len(all_records)} records in Airtable")
        return all_records
    
    def fetch_vesta_loan(self, loan_number: str) -> Optional[Dict]:
        """Fetch loan data from Vesta API by loan number"""
        headers = {
            'Authorization': f'Bearer {VESTA_API_KEY}',
            'Accept': 'application/json',
            'X-Api-Version': VESTA_VERSION
        }
        
        try:
            url = f'{VESTA_BASE_URL}/v1/loans/{loan_number}'
            
            print(f"Fetching Vesta data for loan {loan_number}")
            response = requests.get(url, headers=headers)
            response.raise_for_status()
            return response.json()
            
        except requests.exceptions.HTTPError as e:
            if e.response.status_code == 404:
                self.log_warning(loan_number, 'Vesta API', 'Loan not found in Vesta')
            else:
                self.log_error(loan_number, 'Vesta API', f'HTTP {e.response.status_code}: {str(e)}')
            return None
        except Exception as e:
            self.log_error(loan_number, 'Vesta API', str(e))
            return None
    
    def fetch_parseur_document(self, loan_number: str) -> Optional[Dict]:
        """Fetch parsed document from Parseur by loan number"""
        headers = {
            'Authorization': f'Bearer {PARSEUR_API_KEY}',
            'Content-Type': 'application/json'
        }
        
        try:
            # Get mailbox ID first
            mailboxes_url = 'https://api.parseur.com/v1/mailboxes'
            response = requests.get(mailboxes_url, headers=headers)
            response.raise_for_status()
            
            mailboxes = response.json()
            mailbox_id = None
            
            for mailbox in mailboxes:
                if PARSEUR_MAILBOX_NAME.lower() in mailbox.get('name', '').lower().replace(' ', '-'):
                    mailbox_id = mailbox.get('id')
                    break
            
            if not mailbox_id:
                self.log_error(loan_number, 'Parseur', f'Mailbox "{PARSEUR_MAILBOX_NAME}" not found')
                return None
            
            # Fetch documents from mailbox
            docs_url = f'https://api.parseur.com/v1/mailboxes/{mailbox_id}/documents'
            response = requests.get(docs_url, headers=headers)
            response.raise_for_status()
            
            documents = response.json()
            
            # Find document matching loan number
            for doc in documents:
                # Check if loan number appears in document name or parsed data
                doc_data = doc.get('parsed_data', {})
                doc_name = doc.get('name', '')
                
                if loan_number in doc_name or doc_data.get('document_name') == loan_number:
                    print(f"Found Parseur document for loan {loan_number}")
                    return doc_data
            
            self.log_warning(loan_number, 'Parseur', 'No matching document found')
            return None
            
        except Exception as e:
            self.log_error(loan_number, 'Parseur', str(e))
            return None
    
    def get_county_from_address(self, address: str, city: str, state: str, zip_code: str) -> Optional[str]:
        """Derive county from address information using geocoding"""
        try:
            # Try with just zip code first (most reliable) - restrict to US
            if zip_code:
                location = self.geolocator.geocode(f"{zip_code}, USA", addressdetails=True, country_codes='us')
                if location and 'address' in location.raw:
                    county = location.raw['address'].get('county')
                    if county:
                        county = county.replace(' County', '').strip()
                        return county
            
            # Try with city, state, zip - restrict to US
            if city and state and zip_code:
                fallback_address = f"{city}, {state} {zip_code}, USA"
                location = self.geolocator.geocode(fallback_address, addressdetails=True, country_codes='us')
                if location and 'address' in location.raw:
                    county = location.raw['address'].get('county')
                    if county:
                        county = county.replace(' County', '').strip()
                        return county
            
            # Try with full address - restrict to US
            if address and city and state and zip_code:
                full_address = f"{address}, {city}, {state} {zip_code}, USA"
                location = self.geolocator.geocode(full_address, addressdetails=True, country_codes='us')
                if location and 'address' in location.raw:
                    county = location.raw['address'].get('county')
                    if county:
                        county = county.replace(' County', '').strip()
                        return county
                    
        except (GeocoderTimedOut, GeocoderServiceError) as e:
            print(f"Geocoding error: {e}")
        except Exception as e:
            print(f"Unexpected error in geocoding: {e}")
        
        return None
    
    def process_loan(self, record: Dict) -> Dict:
        """Process a single loan record by combining data from all sources"""
        fields = record.get('fields', {})
        
        # Extract loan number
        loan_number_field = fields.get('Loan Number (from Data Input)')
        if loan_number_field and isinstance(loan_number_field, list) and len(loan_number_field) > 0:
            loan_number = loan_number_field[0]
        else:
            loan_number = fields.get('Name', 'UNKNOWN')
        
        print(f"\n{'='*60}")
        print(f"Processing Loan: {loan_number}")
        print(f"{'='*60}")
        
        borrower_name = fields.get('Name', '')
        
        # Get loan size
        loan_size = fields.get('Loan Size')
        if isinstance(loan_size, list):
            loan_size = loan_size[0] if loan_size else None
        
        # Start building result with Airtable data
        result = {
            'Channel Identifier': 'INDECOMM',
            'Loan Number': loan_number,
            'Loan Amount': loan_size,
            'Borrower Name': borrower_name,
            'Trigger Date': fields.get('Funding Date'),
            'Investor Name': fields.get('Investor'),
        }
        
        # Fetch Vesta data (property information)
        vesta_data = self.fetch_vesta_loan(loan_number)
        if vesta_data:
            # Extract property from subjectProperty in loan response
            subject_property = vesta_data.get('subjectProperty', {})
            address = subject_property.get('address', {})
            
            property_line = address.get('line')
            property_state = address.get('state')
            property_zip = address.get('zipCode')
            property_city = address.get('city')
            
            result['Property Address Line 1'] = property_line
            result['Property State'] = property_state
            result['Property Zip Code'] = property_zip
            
            # Geocode for county since Vesta doesn't provide it
            if property_zip or property_city:
                county = self.get_county_from_address(property_line, property_city, property_state, property_zip)
                if county:
                    result['Property County'] = county
                else:
                    self.log_warning(loan_number, 'Property County', 'Could not derive from geocoding')
            else:
                self.log_warning(loan_number, 'Property County', 'Insufficient address data for geocoding')
            
            if not property_line:
                self.log_warning(loan_number, 'Property Address', 'Not found in Vesta response')
        else:
            self.log_warning(loan_number, 'Property Data', 'No Vesta data available')
        
        # Fetch Parseur data (settlement agent information)
        parseur_data = self.fetch_parseur_document(loan_number)
        if parseur_data:
            result['Organization Name'] = parseur_data.get('settlement_agent')
            result['Organization Phone #'] = parseur_data.get('settlement_phone')
            result['Organization Email'] = parseur_data.get('settlement_agent_email')
        else:
            self.log_warning(loan_number, 'Settlement Agent', 'No Parseur data available')
        
        return result
    
    def generate_excel(self, loan_data: List[Dict], output_path: str):
        """Generate Excel file from loan data"""
        print(f"\nGenerating Excel file...")
        
        if not os.path.exists(TEMPLATE_PATH):
            raise FileNotFoundError(f"Template not found: {TEMPLATE_PATH}")
        
        wb = load_workbook(TEMPLATE_PATH)
        sheet = wb.active
        
        # Column mapping based on template
        column_map = {
            'Channel Identifier': 1,
            'Loan Number': 2,
            'Loan Amount': 5,
            'Borrower Name': 14,
            'Property Address Line 1': 15,
            'Property State': 20,
            'Property Zip Code': 21,
            'Property County': 22,
            'Trigger Date': 24,
            'Organization Name': 34,
            'Organization Phone #': 36,
            'Organization Email': 38,
            'Investor Name': 46,
        }
        
        for row_idx, loan in enumerate(loan_data, start=2):
            for field, col_idx in column_map.items():
                value = loan.get(field)
                if value is not None:
                    sheet.cell(row=row_idx, column=col_idx, value=value)
        
        wb.save(output_path)
        print(f"Excel file saved to: {output_path}")
    
    def generate_processing_notes(self, output_path: str):
        """Generate processing notes file"""
        with open(output_path, 'w') as f:
            f.write("PROCESSING NOTES\n")
            f.write("="*60 + "\n\n")
            
            f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            
            if self.errors:
                f.write(f"ERRORS ({len(self.errors)}):\n")
                f.write("-"*60 + "\n")
                for error in self.errors:
                    f.write(f"  {error}\n")
                f.write("\n")
            else:
                f.write("No errors encountered.\n\n")
            
            if self.warnings:
                f.write(f"WARNINGS ({len(self.warnings)}):\n")
                f.write("-"*60 + "\n")
                for warning in self.warnings:
                    f.write(f"  {warning}\n")
                f.write("\n")
            else:
                f.write("No warnings.\n\n")
        
        print(f"Processing notes saved to: {output_path}")
    
    def run(self):
        """Main execution flow"""
        try:
            print("Starting Indecomm automation...")
            print(f"Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            
            # Fetch records from Airtable
            records = self.fetch_airtable_records()
            
            if not records:
                print("No records found in Airtable view")
                return
            
            # Process each loan
            loan_data = []
            for record in records:
                try:
                    loan_result = self.process_loan(record)
                    loan_data.append(loan_result)
                except Exception as e:
                    loan_number = record.get('fields', {}).get('Name', 'UNKNOWN')
                    self.log_error(loan_number, 'Processing', str(e))
            
            # Generate output files
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            excel_output = f'Funded_File_{timestamp}.xlsx'
            notes_output = f'processing_notes_{timestamp}.txt'
            
            self.generate_excel(loan_data, excel_output)
            self.generate_processing_notes(notes_output)
            
            print(f"\n{'='*60}")
            print("Processing complete!")
            print(f"Processed {len(loan_data)} loans")
            print(f"Errors: {len(self.errors)}")
            print(f"Warnings: {len(self.warnings)}")
            print(f"{'='*60}\n")
            
        except Exception as e:
            print(f"FATAL ERROR: {str(e)}")
            sys.exit(1)

if __name__ == '__main__':
    if not AIRTABLE_TOKEN:
        print("ERROR: AIRTABLE_TOKEN environment variable not set")
        sys.exit(1)
    
    processor = IndecommAutomation()
    processor.run()
